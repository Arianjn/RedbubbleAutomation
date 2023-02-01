import time
from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import TimeoutException
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import pickle

link = "https://www.redbubble.com/portfolio/images/136249712-alisson-becker-in-the-streets-johnny-sins-in-the-sheets/duplicate"

uploadFail = 0
def redbubble_upload(titleVar, descriptionVar, tagVar, imagepathVar):

    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()
    options.add_argument('--user-data-dir=C://Users/Aryan/AppData/Local/Google/Chrome/User Data') #Path to your chrome profile
    options.add_argument("--profile-directory=Default")

    driver = webdriver.Chrome(options=options)

    time.sleep(5)
    driver.get(link)
    time.sleep(3)
    timeout = 10
    try: # Security Checks and shit
        element_present = expected_conditions.presence_of_element_located((By.ID, "work_title_en"))
        WebDriverWait(driver, timeout).until(element_present)
    except TimeoutException:
        print("Timed out waiting for page to load")
    time.sleep(3)
    #Input Title
    driver.find_element(By.ID, "work_title_en").clear()
    time.sleep(3)
    driver.find_element(By.ID, "work_title_en").send_keys(titleVar)

    #Input Tags
    time.sleep(3)
    driver.find_element(By.ID, "work_tag_field_en").clear()
    time.sleep(3)
    driver.find_element(By.ID, "work_tag_field_en").send_keys(tagVar)

    #Input Description
    time.sleep(3)
    driver.find_element(By.ID, "work_description_en").clear()
    time.sleep(3)
    driver.find_element(By.ID, "work_description_en").send_keys(descriptionVar)
    time.sleep(3)

    #InsertImage
    try:
        driver.find_element(By.ID, "select-image-base").send_keys(imagepathVar)
    except:
        uploadFail = 1

    time.sleep(30)

    driver.find_element(By.ID,"rightsDeclaration").click()



    driver.find_element(By.ID,"submit-work").click()

    time.sleep(40)
    # #CenterImage
    # driver.find_element(By.XPATH,"(//div[text()='Edit'])[4]").click()
    #
    # time.sleep(4)
    # element = driver.find_element(By.XPATH,"(//button[text()='Center horizontally'")
    # actions = ActionChains(driver)
    # actions.move_to_element(element).perform()


    driver.quit()

#Redbubble Upload-----------------------------------------------------------------------RedbubbleUpload


wb = load_workbook("player_name.xlsx")
ws = wb.active

titlePrefix = "in the Streets, Johnny Sins in the Sheets"
imgPath = r"C:\Users\Aryan\Documents\Redbubble_Designs\PlayerTshirts\\"
tagsPrefix = ", meme, Johnny Sins, street football, football, soccer, funny, witty, hilarious"
descriptionPrefix = "in the Streets, Johnny Sins in the Sheets based on the in the streets vs in the sheets meme"
#imageDash =
imageNumber = -1
#rowNumber = 0
savedImageNumber = 157

customNumberWanted = int(input("do you want to start at custom point? "))

if(customNumberWanted ==1):
    customsavedImageNumber = int(input("Custom Point: "))
    with open("imagenumber.txt", "wb") as file:
        pickle.dump(customsavedImageNumber, file)
        file.close()
for cell in ws['A']:

    #rowNumber = str(cell[-2])
    imageNumber += 1


    with open("imagenumber.txt", "rb") as theFile:
        savedImageNumber = pickle.load(theFile)

    if str(cell) == "<Cell 'player_name'.A1>":
        pass
    elif len(cell.value) > 15:
        pass
    elif imageNumber <= savedImageNumber:
        pass
    elif uploadFail == 1:
        pass
        uploadFail = 0
    else:
        title = str.title(cell.value) + " " + titlePrefix
        image = imgPath + cell.value + str(imageNumber) +".png"
        tags = str.title(cell.value) + tagsPrefix
        description = str.title(cell.value) + " " + descriptionPrefix
        #print(image)
        with open("imagenumber.txt", "wb") as file:
            pickle.dump(imageNumber, file)
            file.close()
        print(imageNumber)
        redbubble_upload(title, description, tags, image)


